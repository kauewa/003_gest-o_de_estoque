from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def dia_de_hoje(formato='%d/%m/%Y'):
    data_atual = datetime.now()
    return data_atual.strftime(formato)


def preencher_planilha_venda(dados:list, nome:str):
    wb = load_workbook('modelo_planilha_venda.xlsx')
    ws = wb.active  # Selecione a primeira aba da planilha
    img = Image('Imagem_planilha.png')
    total = 0
    # Suponha que 'dados' seja uma lista de listas, onde cada lista representa uma linha de dados
    for i, data in enumerate(dados, start=22): 
        ws.cell(row=i, column=1, value=data['COD_P'])
        ws.cell(row=i, column=2, value=data['PRODUTO'])
        ws.cell(row=i, column=3, value=data['PESO E LÍQUIDO'])
        ws.cell(row=i, column=4, value=data['FARDOS'])
        ws.cell(row=i, column=5, value=data['UNIDADE POR FARDOS'])
        ws.cell(row=i, column=6, value=data['VALOR UNITARIO'])
        ws.cell(row=i, column=7, value=data['VALOR POR FARDO'])
        total += float(data['FARDOS']) * data['VALOR POR FARDO']
    # Salve a planilha em um objeto BytesIO
    ws['C18'] = nome
    ws.add_image(img, 'H25')
    ws['H24'] = total

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def float_para_dinheiro(valor):
    return 'R$ {:,.2f}'.format(valor).replace('.', '!').replace(',', '.').replace('!', ',')